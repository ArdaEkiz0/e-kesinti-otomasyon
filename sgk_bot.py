"""SGK Bot - Sistem Fonksiyonlarini Tetikleyen Versiyon"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementNotInteractableException,
    ElementClickInterceptedException,
)
import pandas as pd
from datetime import datetime, timedelta
import time, os, sys, re, json, urllib.request, urllib.error, subprocess

# Uygulama sürümü ve güncelleme kontrolü
BOT_SURUM = "1.7.20"  # GitHub release etiketiyle karsilastirilir
GITHUB_REPO = "ArdaEkiz0/e-kesinti-otomasyon"
GITHUB_API = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"

# Türkçe karakterler ve emojiler hangi konsolda olursa olsun yazılabilsin
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# Konsol kod sayfasını UTF-8'e çevir (Türkçe karakterler bozulmasın)
if os.name == "nt":
    try:
        import ctypes
        ctypes.windll.kernel32.SetConsoleOutputCP(65001)
        ctypes.windll.kernel32.SetConsoleCP(65001)
    except Exception:
        pass

# Cmd pencere başlığını ayarla (sürüm bilgisiyle birlikte)
if os.name == "nt":
    try:
        import ctypes
        ctypes.windll.kernel32.SetConsoleTitleW(f"SGK E-Kesinti Otomasyon v{BOT_SURUM} | Developer Arda M. Ekiz")
    except Exception:
        pass

# ============================================================
# RENKLİ ÇIKTI YARDIMCILARI (Kullanıcı dostu arayüz)
# ============================================================

class Renk:
    RESET   = "\033[0m"
    BOLD    = "\033[1m"
    KIRMIZI = "\033[31m"
    YESIL   = "\033[32m"
    SARI    = "\033[33m"
    MAVI    = "\033[34m"
    MOR     = "\033[35m"
    TURKUAZ = "\033[36m"
    ALTIN   = "\033[38;2;201;168;108m"  # marka altini (#c9a86c)


def renk_aktif_mi():
    """Renklerin açık olup olmadığını belirler (SGK_RENK=1 zorlar, 0 kapatır)."""
    zorla = os.environ.get("SGK_RENK")
    if zorla == "1":
        return True
    if zorla == "0":
        return False
    if os.name == "nt":
        os.system("")  # Windows 10+ konsolda ANSI renklerini aktifleştirir
    return sys.stdout.isatty()


RENK_ACIK = renk_aktif_mi()


def renkli(metin, renk=Renk.RESET, kalin=False):
    """Metni istenen renk ve kalınlıkta döndürür (renk kapalıysa düz metin)."""
    if not RENK_ACIK:
        return str(metin)
    oz = (Renk.BOLD if kalin else "") + renk
    return f"{oz}{metin}{Renk.RESET}"


def ilerleme_cubugu(mevcut, toplam, genislik=20):
    """Örn: [████████░░░░░░░░░░░░]  40% (6/14)"""
    if toplam <= 0:
        return ""
    yuzde = mevcut / toplam * 100
    dolu = int(genislik * mevcut / toplam)
    cubuk = "█" * dolu + "░" * (genislik - dolu)
    return renkli(f"[{cubuk}] {yuzde:3.0f}% ({mevcut}/{toplam})", Renk.TURKUAZ)


def hata_aciklamasi(e):
    """Selenium hatalarını kullanıcının anlayacağı Türkçe mesaja çevirir."""
    sinif = type(e).__name__
    metin = str(e).lower()
    if isinstance(e, NoSuchElementException) or "no such element" in metin:
        return "Sayfa elemanı bulunamadı (SGK sayfası değişmiş olabilir)"
    if isinstance(e, TimeoutException) or "timeout" in metin:
        return "Zaman aşımı (sayfa yavaş yükleniyor olabilir)"
    if isinstance(e, ElementNotInteractableException) or "not interactable" in metin:
        return "Elemanla etkileşim kurulamıyor (gizli veya kapalı olabilir)"
    if isinstance(e, ElementClickInterceptedException) or "intercepted" in metin:
        return "Tıklama engellendi (elemanın üstünde başka eleman var)"
    if "alert" in metin:
        return "Sayfada açılır pencere (alert) bekleniyor, elle kapatın"
    return (metin[:200] if metin else f"Bilinmeyen hata ({sinif})")


def guncelleme_var_mi():
    """GitHub'da daha yeni surum varsa (surum, zip_url) doner; yoksa (None, None)."""
    try:
        istek = urllib.request.Request(
            GITHUB_API, headers={"User-Agent": "SGK-BOT", "Accept": "application/vnd.github+json"})
        with urllib.request.urlopen(istek, timeout=8) as r:
            veri = json.loads(r.read().decode("utf-8"))
        uzak_surum = str(veri.get("tag_name", "")).lstrip("v")
        if not uzak_surum:
            return None, None
        yerel = [int(x) for x in re.findall(r"\d+", BOT_SURUM)]
        uzak = [int(x) for x in re.findall(r"\d+", uzak_surum)]
        uzunluk = max(len(yerel), len(uzak))
        yerel += [0] * (uzunluk - len(yerel))
        uzak += [0] * (uzunluk - len(uzak))
        if uzak <= yerel:
            return None, None
        for a in veri.get("assets", []):
            if a.get("name", "").lower() == "sgk_e_kesinti_otomasyon.zip":
                return uzak_surum, a.get("browser_download_url")
    except Exception:
        pass  # internet yoksa veya hata olursa sessizce gec
    return None, None


def otomatik_guncelle(zip_url):
    """Yeni ZIP paketini indirip bot klasorune acar (dosyalari gunceller)."""
    import zipfile
    hedef_klasor = os.path.dirname(os.path.abspath(__file__))
    zip_yol = os.path.join(hedef_klasor, "_guncelleme.zip")
    yaz_ilerleme = True
    print(renkli("   ⬇️ Yeni sürüm indiriliyor...", Renk.SARI))
    istek = urllib.request.Request(zip_url, headers={"User-Agent": "SGK-BOT"})
    with urllib.request.urlopen(istek, timeout=180) as r, open(zip_yol, "wb") as f:
        toplam = int(r.headers.get("Content-Length", 0))
        yazilan = 0
        while True:
            parca = r.read(65536)
            if not parca:
                break
            f.write(parca)
            yazilan += len(parca)
            if toplam:
                print(f"\r   %3d%%" % (yazilan * 100 // toplam), end="", flush=True)
    print()
    if os.path.getsize(zip_yol) < 1000:
        raise RuntimeError("ZIP indirilemedi (dosya eksik)")
    print(renkli("   📦 Dosyalar güncelleniyor...", Renk.SARI))
    with zipfile.ZipFile(zip_yol) as z:
        z.extractall(hedef_klasor)
    try:
        os.remove(zip_yol)
    except OSError:
        pass


def windows_bildirim(baslik, metin):
    """İşlem sonunda Windows balon bildirimi gösterir (hata olursa sessizce geçer)."""
    if os.name != "nt":
        return
    try:
        import subprocess
        def _ps_escape(s):
            """PowerShell string kaçış: tek tırnak ve özel karakterleri temizler."""
            return str(s).replace("'", "''").replace("`", "").replace("$", "").replace(";", "")
        ps = (
            "Add-Type -AssemblyName System.Windows.Forms;"
            "Add-Type -AssemblyName System.Drawing;"
            "$n = New-Object System.Windows.Forms.NotifyIcon;"
            "$n.Icon = [System.Drawing.SystemIcons]::Information;"
            f"$n.BalloonTipTitle = '{_ps_escape(baslik)}';"
            f"$n.BalloonTipText = '{_ps_escape(metin)}';"
            "$n.Visible = $true;"
            "$n.ShowBalloonTip(8000)"
        )
        subprocess.run(["powershell", "-NoProfile", "-Command", ps],
                       capture_output=True, timeout=20)
    except Exception:
        pass


# ============================================================
# TEST MODU İÇİN SAHTE SÜRÜCÜ (tarayıcı açmadan arayüzü dener)
# ============================================================

class TestEleman:
    def __init__(self, surucu, kimlik):
        self.surucu = surucu
        self.kimlik = kimlik

    def clear(self):
        pass

    def send_keys(self, deger):
        if self.kimlik == "mernisno":
            self.surucu.son_tc = str(deger)

    def click(self):
        pass


class TestSurucu:
    """Gerçek Chrome açmadan bot akışını simüle eden sahte sürücü."""
    def __init__(self, hatali_tcler=()):
        self.hatali = set(str(tc) for tc in hatali_tcler)
        self.son_tc = None

    def get(self, url):
        pass

    def find_element(self, by, deger):
        return TestEleman(self, deger)

    def execute_script(self, script):
        if self.son_tc in self.hatali:
            raise Exception(f"Simüle hata (TC={self.son_tc})")
        return True


# ============================================================
# ANA BOT SINIFI
# ============================================================

MAX_DENEME = 3          # hata olursa her kayit en fazla 3 kez denenir
MAKS_YENIDEN_LOGIN = 5  # tek turda en fazla kac kez oturum yenileme istenir


class OturumHatasi(Exception):
    """SGK oturumu dusmustu / login sayfasina yonlendirildi."""


class SGKBot:
    def __init__(self, surucu=None, test_modu=False):
        self.test_modu = test_modu
        self.kooperatif_adi = None
        if surucu is not None:
            self.driver = surucu
        else:
            self.driver = self._start_driver()

    def _temizle_eski_kilitler(self, en_fazla_dk=5):
        """webdriver-manager'ın yarıda kalan indirmelerden bıraktığı eski kilit dosyalarını siler."""
        wdm = os.path.expanduser("~/.wdm")
        if not os.path.isdir(wdm):
            return
        simdi = time.time()
        for kilit in os.listdir(wdm):
            if not kilit.startswith(".wdm-lock-"):
                continue
            yol = os.path.join(wdm, kilit)
            try:
                if simdi - os.path.getmtime(yol) > en_fazla_dk * 60:
                    os.remove(yol)
                    print(f"   🧹 Eski sürücü kilidi temizlendi: {kilit}")
            except OSError:
                pass

    def _start_driver(self):
        """Chrome sürücüsünü başlatır. webdriver-manager başarısız olursa yerel chromedriver.exe'yi dener."""
        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option("useAutomationExtension", False)
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        self._temizle_eski_kilitler()
        son_hata = None
        for deneme in range(2):
            try:
                driver = webdriver.Chrome(
                    service=Service(ChromeDriverManager().install()),
                    options=chrome_options
                )
                driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                    "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"
                })
                return driver
            except Exception as e:
                son_hata = e
                if "lock" in str(e).lower():
                    # Yarıda kalan indirme kilidi kalmış: hepsini temizle, bir kez daha dene
                    self._temizle_eski_kilitler(en_fazla_dk=0)
        yerel = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chromedriver.exe")
        if os.path.exists(yerel):
            try:
                driver = webdriver.Chrome(service=Service(yerel), options=chrome_options)
                driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                    "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"
                })
                return driver
            except Exception as e2:
                son_hata = e2
        raise RuntimeError(
            "Chrome tarayıcısı başlatılamadı!\n"
            "   • İnternet bağlantınızı kontrol edin (chromedriver indirilemiyor olabilir)\n"
            "   • Chrome sürümünüzle uyumlu chromedriver gerekiyor (Chrome → Ayarlar → Hakkında)\n"
            f"   • Teknik detay: {str(son_hata)[:200]}"
        )

    def get_previous_month_date(self):
        today = datetime.now()
        first_day = today.replace(day=1)
        last_day = first_day - timedelta(days=1)
        return last_day.strftime("%d.%m.%Y")

    def get_previous_month_and_year(self):
        today = datetime.now()
        first_day = today.replace(day=1)
        last_day = first_day - timedelta(days=1)
        return last_day.month, last_day.year

    @staticmethod
    def _kooperatif_adi_oku(file_path):
        """Excel'in ilk satırındaki (A1) kooperatif adını okur. Bulunamazsa None döner."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ilk_sayfa = wb[wb.sheetnames[0]]
            deger = ilk_sayfa.cell(row=1, column=1).value
            wb.close()
            if deger and str(deger).strip():
                ad = str(deger).strip()
                # Şablon başlığı gibi görünen değerleri kooperatif adı sanma
                if ad not in ("Ünvan", "Unvan", "TC Kimlik No", "Matrah", "Bağ-Kur", "Bağkur"):
                    return ad
        except Exception:
            pass
        return None

    def load_excel_data(self, file_path):
        try:
            self.kooperatif_adi = self._kooperatif_adi_oku(file_path)
            df = pd.read_excel(file_path)
            data = []
            bolunen_sayisi = 0
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[1]) and len(str(row.iloc[1])) == 11:
                    try:
                        tc_no = str(int(row.iloc[1]))
                        matrah = self._sayiya_cevir(row.iloc[2])
                        kesinti = self._sayiya_cevir(row.iloc[3])
                        # Matrah 1 milyon ve üzeriyse ikiye böl
                        if matrah >= 1_000_000:
                            yari_matrah = round(matrah / 2, 2)
                            yari_kesinti = round(kesinti / 2, 2)
                            data.append({'tc': tc_no, 'matrah': yari_matrah, 'kesinti': yari_kesinti})
                            data.append({'tc': tc_no, 'matrah': yari_matrah, 'kesinti': yari_kesinti})
                            bolunen_sayisi += 1
                        else:
                            data.append({'tc': tc_no, 'matrah': matrah, 'kesinti': kesinti})
                    except Exception:
                        continue
            if self.kooperatif_adi:
                print(f"   🏢 Kooperatif: {renkli(self.kooperatif_adi, Renk.SARI, kalin=True)}")
            print(f"✅ Excel'den {len(data)} kayıt okundu", end="")
            if bolunen_sayisi > 0:
                print(f" ({renkli(str(bolunen_sayisi) + ' kayıt bölündü', Renk.SARI)})")
            else:
                print()
            return data
        except Exception as e:
            print(f"❌ Excel okunamadı: {e}")
            return []

    @staticmethod
    def _sayiya_cevir(deger):
        """Excel değerini sayıya çevirir. '13014,09' gibi virgüllü metinleri de destekler."""
        if isinstance(deger, str):
            return float(deger.replace(",", "."))
        return float(deger)

    @staticmethod
    def excel_sablonu_hazirla(dosya_yolu="çalışmaaaa.xlsx"):
        """Excel dosyası yoksa standart şablonu oluşturur. Mevcut dosyaya dokunmaz."""
        import openpyxl
        from openpyxl.styles import Font
        if os.path.exists(dosya_yolu):
            return False
        simdi = datetime.now()
        son_gun = (simdi.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kesinti"
        ws["A1"] = "S.S.AYDOĞDU KÖYÜ TAR.KALK.KOOP."
        ws["A2"] = f"Tarih: 01.{simdi.month:02d}.{simdi.year} -> {son_gun.day:02d}.{simdi.month:02d}.{simdi.year}"
        ws["A3"] = "Total sahaların hepsi sıfır ise alınmasın"
        ws["A4"] = "Toplu Alım Makbuz Raporu_bağkur_tevk"
        for i, baslik in enumerate(["Ünvan", "TC Kimlik No", "Matrah", "Bağ-Kur"], start=1):
            hucre = ws.cell(row=5, column=i, value=baslik)
            hucre.font = Font(bold=True)
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        wb.save(dosya_yolu)
        return True

    def _form_sayfasina_git(self):
        """Login sonrasi islem formuna geri doner (oturum yenilendikten sonra)."""
        print(renkli("\n📍 Form sayfasına gidiliyor...", Renk.TURKUAZ))
        try:
            self.driver.execute_script("goToPage('/GooKesintiIcinUreticiSorgulaAction.do')")
            time.sleep(3)
        except Exception:
            pass
        try:
            self.driver.execute_script("acikDosyaSayisiniKontrolEt()")
            time.sleep(3)
        except Exception:
            pass
        month, year = self.get_previous_month_and_year()
        try:
            Select(self.driver.find_element(By.ID, "ay")).select_by_value(str(month))
            Select(self.driver.find_element(By.ID, "yil")).select_by_value(str(year))
            print(f"   ✅ Ay: {month:02d}, Yıl: {year}")
        except Exception as e:
            print(renkli(f"   ❌ Ay/Yıl hatası: {e}", Renk.KIRMIZI))
        try:
            self.driver.execute_script("document.getElementById('yeniDosyaOlustur').click()")
            time.sleep(4)
            print(renkli("   ✅ YENİ Dosya oluşturuldu", Renk.YESIL))
        except Exception as e:
            print(renkli(f"   ❌ Hata: {e}", Renk.KIRMIZI))
        try:
            time.sleep(2)
            rows = self.driver.find_elements(By.XPATH, "//table[@id='users2']//tbody//tr[@class='row1' or @class='row2']")
            if len(rows) > 0:
                rows[0].click()
                time.sleep(1)
                print(renkli("   ✅ YENİ Dosya seçildi", Renk.YESIL))
        except Exception:
            pass
        button = self.driver.find_element(By.ID, "secilenDosyayiAc")
        button.click()
        time.sleep(3)
        print(renkli("   ✅ Form açıldı - işleme devam ediliyor...", Renk.YESIL))

    def _kayit_isle(self, record, odeme_tarihi):
        """Tek bir kaydi siteye girer. Basarili olursa doner, hata olursa exception firlatir."""
        try:
            tc_field = self.driver.find_element(By.ID, "mernisno")
        except NoSuchElementException:
            # TC alani yok -> oturum dussm ya da baska sayfaya yonlendirilmis olabilir
            try:
                url = (self.driver.current_url or "").lower()
            except Exception:
                url = ""
            if "sgk.gov.tr" not in url:
                raise OturumHatasi("SGK oturumu dusmus / login gerekiyor")
            raise
        tc_field.clear()
        tc_field.send_keys(record['tc'])
        time.sleep(0.5)
        print(renkli(f"   ✅ TC yazıldı: {record['tc']}", Renk.YESIL))

        tarih_field = self.driver.find_element(By.ID, "odemeTarihi")
        tarih_field.clear()
        tarih_field.send_keys(odeme_tarihi)
        time.sleep(0.5)
        print(renkli(f"   ✅ Tarih yazıldı: {odeme_tarihi}", Renk.YESIL))

        self.driver.execute_script("""
            var radio = document.querySelector('input[name="answer"][value="Alim Gerceklesti"]');
            if (radio) {
                var link = radio.parentElement.querySelector('a');
                if (link) {
                    link.click();
                }
            }
        """)
        time.sleep(0.5)
        print(renkli("   ✅ 'Alım Gerceklesti' seçildi (YEŞİL)", Renk.YESIL))

        print(renkli("   🔎 Ara tıklanıyor...", Renk.TURKUAZ))
        self.driver.execute_script("tcAraButtonPressed()")
        time.sleep(3)
        print(renkli("   ✅ Üretici bilgileri yüklendi", Renk.YESIL))

        matrah_tam = int(record['matrah'])
        matrah_kurus = round((record['matrah'] - matrah_tam) * 100)
        # Kuruş HER ZAMAN 2 haneli yazılır (ör: 9 -> '09', 5 -> '05', 90 -> '90')
        # '9' yazılırsa site bunu 0.9 (13014,90) olarak algılar!
        matrah_kurus_str = f"{matrah_kurus:02d}"

        # Matrah yazı - SİSTEM FONKSİYONLARINI TETIKLE
        self.driver.execute_script(f"""
            var input = document.kesinti.urunAlimBedeliTam;
            input.value = '{matrah_tam}';
            textFocus(input);
            formatCurrency(input);
            hesaplaYaz(input);
        """)
        time.sleep(0.5)

        # Kuruş yazılır; sitenin formatCurrency/hesaplaYaz fonksiyonları
        # baştaki sıfırı silebileceği için değer EN SONDA tekrar '09' olarak
        # yazılır ve hesaplama bir kez daha tetiklenir.
        self.driver.execute_script(f"""
            var input = document.kesinti.urunAlimBedeliKrs;
            input.value = '{matrah_kurus_str}';
            textFocus(input);
            formatCurrency(input);
            hesaplaYaz(input);
            input.value = '{matrah_kurus_str}';
            hesaplaYaz(input);
        """)
        time.sleep(0.5)
        print(renkli(f"   💰 Matrah yazıldı: {matrah_tam}.{matrah_kurus_str} (RENK DEĞİŞTİ)", Renk.YESIL))

        print(f"   🔢 Kesinti Tutarı: {record['kesinti']}")

        print(renkli("   ➕ EKLE tıklanıyor...", Renk.TURKUAZ))
        self.driver.execute_script("kesintiKaydet('9999999999.00')")
        time.sleep(2)

    def _toplu_isle(self, kayitlar, odeme_tarihi):
        """Verilen kayitlari sirayla isler. Dondugunde (basarili_sayisi, hatali_listesi).

        Oturum duserse kullanicidan tekrar login istenip kaldigi yerden devam edilir;
        boylece 200+ kayitlik listelerde oturum kaybi tum listeyi felc etmez."""
        basarili_sayi = 0
        hatali = []
        toplam = len(kayitlar)
        tahmin_sn = toplam * 9  # kayit basi ~9 saniye (site bekleme sureleri dahil)
        print(renkli(f"   ⏱️  Tahmini süre: ~{tahmin_sn // 60} dk {tahmin_sn % 60} sn", Renk.TURKUAZ))

        tur_baslangici = time.time()
        try:
            for idx, record in enumerate(kayitlar, 1):
                gecen = time.time() - tur_baslangici
                kalan = gecen / idx * (toplam - idx) if idx > 1 else tahmin_sn
                print(renkli(f"\n📍 Kayıt {idx}/{toplam}", Renk.BOLD + Renk.MAVI, kalin=True))
                print(f"   {ilerleme_cubugu(idx - 1, toplam)}  TC={record['tc']}  "
                      f"(kalan ~{int(kalan // 60)} dk {int(kalan % 60)} sn)")
                basarili = False
                son_hata = None
                deneme = 0
                girisim = 0
                while deneme < MAX_DENEME and girisim < MAX_DENEME + MAKS_YENIDEN_LOGIN:
                    girisim += 1
                    try:
                        self._kayit_isle(record, odeme_tarihi)
                        basarili = True
                        break
                    except OturumHatasi:
                        if self.test_modu:
                            son_hata = OturumHatasi("test")
                            deneme += 1
                            continue
                        print(renkli("   🔒 SGK oturumu düşmüş görünüyor!", Renk.SARI, kalin=True))
                        try:
                            input(renkli("   Tarayıcıdan TEKRAR LOGIN yapın, sonra buraya dönüp ENTER'a basınız...", Renk.SARI))
                        except EOFError:
                            pass
                        try:
                            self._form_sayfasina_git()
                            continue  # ayni kaydi tekrar dene, deneme hakkini yakma
                        except Exception as e_form:
                            son_hata = e_form
                            deneme += 1
                    except Exception as e:
                        son_hata = e
                        deneme += 1
                        if deneme < MAX_DENEME:
                            print(renkli(f"   ⚠️ Hata ({deneme}. deneme): {hata_aciklamasi(e)}", Renk.SARI))
                            print(renkli(f"   🔄 {deneme + 1}. deneme yapılıyor...", Renk.TURKUAZ))
                            time.sleep(2)
                if basarili:
                    print(renkli("   ✅ TC başarıyla işlendi!", Renk.YESIL, kalin=True))
                    basarili_sayi += 1
                else:
                    mesaj = hata_aciklamasi(son_hata) if son_hata else "Bilinmeyen hata"
                    print(renkli(f"   ❌ Hata ({MAX_DENEME} deneme): {mesaj}", Renk.KIRMIZI))
                    hatali.append({'tc': record['tc'], 'matrah': record['matrah'],
                                   'kesinti': record['kesinti'], 'hata': mesaj})
        except KeyboardInterrupt:
            print(renkli("\n⏸️  Kullanıcı durdurdu — şu ana kadarki özet aşağıda.", Renk.SARI, kalin=True))
        return basarili_sayi, hatali

    def run(self, excel_file):
        baslangic = time.time()
        print("\n" + "="*60)
        print(renkli(r"""
  ███████╗ ██████╗ ██╗  ██╗
  ██╔════╝██╔════╝ ██║ ██╔╝
  ███████╗██║  ███╗█████╔╝
  ╚════██║██║   ██║██╔═██╗
  ███████║╚██████╔╝██║  ██╗
  ╚══════╝ ╚═════╝ ╚═╝  ╚═╝""", Renk.ALTIN, kalin=True))
        print(renkli("   SGK TARIMSAL KESİNTİ OTOMASYONU BOT", Renk.MAVI, kalin=True))
        print(renkli(f"   Sürüm: v{BOT_SURUM}", Renk.TURKUAZ, kalin=True))
        print(renkli("   Developer: Arda M. Ekiz", Renk.MOR))
        print("="*60)

        print(renkli("\n📊 Excel dosyası okunuyor...", Renk.TURKUAZ))
        if self.excel_sablonu_hazirla(excel_file):
            print(renkli(f"   ✅ {os.path.basename(excel_file)} bulunamadı, standart şablon oluşturuldu!", Renk.YESIL))
            print(renkli("   ⚠️  Lütfen şablonu doldurun (Ünvan, TC Kimlik No, Matrah, Bağ-Kur) ve tekrar çalıştırın.", Renk.SARI))
            return
        data = self.load_excel_data(excel_file)
        if not data:
            print(renkli("❌ Excel'de veri yok!", Renk.KIRMIZI, kalin=True))
            return

        if self.test_modu:
            print(renkli("\n🧪 TEST MODU: Tarayıcı açılmadan akış simüle ediliyor...", Renk.SARI, kalin=True))
        else:
            print(renkli("\n🌐 SGK sitesine gidiliyor...", Renk.TURKUAZ))
            self.driver.get("https://uyg.sgk.gov.tr/TAKEP/Welcome.do")
            print(renkli("⏳ Lütfen siteye login yapınız...", Renk.SARI))
            input("   Login yaptıktan sonra ENTER'a basınız...")

            print(renkli("\n📍 Form sayfasına gidiliyor...", Renk.TURKUAZ))
            try:
                self.driver.execute_script("goToPage('/GooKesintiIcinUreticiSorgulaAction.do')")
                time.sleep(3)
            except Exception:
                print(renkli("⚠️  Navigasyon hatası", Renk.SARI))

            print(renkli("\n📂 'Dosya Aç / Çağır' dialog açılıyor...", Renk.TURKUAZ))
            try:
                self.driver.execute_script("acikDosyaSayisiniKontrolEt()")
                time.sleep(3)
            except Exception:
                print(renkli("⚠️  Dialog açılmadı", Renk.SARI))

            print(renkli("\n📅 Ay ve Yıl seçiliyor...", Renk.TURKUAZ))
            month, year = self.get_previous_month_and_year()
            try:
                Select(self.driver.find_element(By.ID, "ay")).select_by_value(str(month))
                Select(self.driver.find_element(By.ID, "yil")).select_by_value(str(year))
                print(f"   ✅ Ay: {month:02d}, Yıl: {year}")
            except Exception as e:
                print(renkli(f"❌ Ay/Yıl hatası: {e}", Renk.KIRMIZI))

            print(renkli("\n📄 'Dosya Oluştur' tıklanıyor...", Renk.TURKUAZ))
            try:
                self.driver.execute_script("document.getElementById('yeniDosyaOlustur').click()")
                time.sleep(4)
                print(renkli("   ✅ YENİ Dosya oluşturuldu", Renk.YESIL))
            except Exception as e:
                print(renkli(f"   ❌ Hata: {e}", Renk.KIRMIZI))

            print(renkli("\n📋 Oluşturulan YENİ dosya seçiliyor...", Renk.TURKUAZ))
            try:
                time.sleep(2)
                rows = self.driver.find_elements(By.XPATH, "//table[@id='users2']//tbody//tr[@class='row1' or @class='row2']")
                print(f"   Bulunan dosya: {len(rows)}")

                if len(rows) > 0:
                    first_row = rows[0]
                    first_row.click()
                    time.sleep(1)
                    print(renkli("   ✅ YENİ Dosya seçildi (SARI)", Renk.YESIL))
            except Exception as e:
                print(renkli(f"   ❌ Hata: {e}", Renk.KIRMIZI))

            print(renkli("\n📂 'Seçilen Dosyayı Aç' tıklanıyor...", Renk.TURKUAZ))
            try:
                time.sleep(1)
                button = self.driver.find_element(By.ID, "secilenDosyayiAc")
                button.click()
                time.sleep(3)
                print(renkli("   ✅ Form açıldı - TC işlemeye başlanıyor...", Renk.YESIL))
            except Exception as e:
                print(renkli(f"   ❌ Hata: {e}", Renk.KIRMIZI))

        print(renkli("\n📝 Kayıtlar işleniyor...", Renk.TURKUAZ, kalin=True))
        odeme_tarihi = self.get_previous_month_date()
        success, hatali = self._toplu_isle(data, odeme_tarihi)

        # Hatali kayitlar icin bir kez daha deneme turu (oturum dususu gibi gecici
        # hatalardan kaynaklanan kayitlar bu sayede kurtarilir)
        if hatali and not self.test_modu:
            try:
                cevap = input(renkli(
                    f"\n🔁 {len(hatali)} kayıt hatalı. Bir kez daha denenmesini ister misiniz? (E/H): ",
                    Renk.SARI, kalin=True)).strip().lower()
            except EOFError:
                cevap = "h"
            if cevap.startswith("e"):
                print(renkli("\n🔄 Hatalı kayıtlar tekrar deneniyor...", Renk.SARI, kalin=True))
                s2, hatali = self._toplu_isle(hatali, odeme_tarihi)
                success += s2

        sure = time.time() - baslangic
        print("\n" + "="*60)
        print(renkli("✅ İŞLEM TAMAMLANDI!", Renk.YESIL, kalin=True))
        print(f"   Başarılı: {renkli(f'{success}/{len(data)}', Renk.YESIL, kalin=True)}"
              f"   Süre: {int(sure // 60)} dk {int(sure % 60)} sn")
        if hatali:
            print(f"   Başarısız: {renkli(str(len(hatali)), Renk.KIRMIZI, kalin=True)}")
            print(renkli("\n   ❌ HATA ALAN KAYITLAR:", Renk.KIRMIZI, kalin=True))
            for h in hatali:
                print(f"     • TC={h['tc']}: {renkli(h['hata'], Renk.KIRMIZI)}")
        else:
            print(renkli("   🎉 Tüm kayıtlar başarıyla işlendi!", Renk.YESIL))
        print("="*60)
        print(renkli("Developer: Arda M. Ekiz", Renk.MOR, kalin=True))

        if not self.test_modu:
            ozet = f"{success}/{len(data)} kayıt başarılı" + (f", {len(hatali)} hatalı" if hatali else ", tümü başarılı!")
            windows_bildirim("SGK Bot - İşlem Tamamlandı", ozet)

        if not self.test_modu:
            input("\nKapatmak için ENTER'a basınız...")


def excel_dosyasi_sec():
    """Klasördeki Excel dosyasını otomatik bulur.
    Tek dosya varsa adı ne olursa olsun onu seçer,
    birden fazlaysa kullanıcıya seçtirir, hiç yoksa şablon adını döner."""
    klasor = os.path.dirname(os.path.abspath(__file__))
    varsayilan = os.path.join(klasor, "çalışmaaaa.xlsx")
    dosyalar = sorted(
        f for f in os.listdir(klasor)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    )
    if not dosyalar:
        return varsayilan  # yoksa şablon otomatik oluşturulur
    if len(dosyalar) == 1:
        return os.path.join(klasor, dosyalar[0])
    # Birden fazla dosya varsa menü göster
    print(renkli("\n📂 Birden fazla Excel dosyası bulundu:", Renk.TURKUAZ, kalin=True))
    varsayilan_idx = 1
    for i, f in enumerate(dosyalar, 1):
        if f == "çalışmaaaa.xlsx":
            varsayilan_idx = i
        isaret = "  (varsayılan)" if f == "çalışmaaaa.xlsx" else ""
        print(f"   [{i}] {f}{isaret}")
    while True:
        secim = input(f"   Hangi dosya işlensin? (1-{len(dosyalar)}, ENTER={varsayilan_idx}): ").strip()
        if secim == "":
            return os.path.join(klasor, dosyalar[varsayilan_idx - 1])
        if secim.isdigit() and 1 <= int(secim) <= len(dosyalar):
            return os.path.join(klasor, dosyalar[int(secim) - 1])
        print(renkli("   ❌ Geçersiz seçim, tekrar deneyin.", Renk.KIRMIZI))


if __name__ == "__main__":
    # Otomatik guncelleme: yeni surum varsa sorar, onaylanirsa indirip uygular
    if "--test" not in sys.argv:
        yeni_surum, zip_url = guncelleme_var_mi()
        if yeni_surum and zip_url:
            print(renkli(f"\n🔄 YENİ SÜRÜM VAR: v{yeni_surum} (senin sürüm: v{BOT_SURUM})", Renk.SARI, kalin=True))
            try:
                cevap = input(renkli("   Şimdi otomatik güncellensin mi? (E/H): ", Renk.SARI)).strip().lower()
            except EOFError:
                cevap = "h"
            if cevap in ("e", "evet", ""):
                try:
                    otomatik_guncelle(zip_url)
                    print(renkli(f"   ✅ v{yeni_surum} güncellendi! Bot yeniden başlatılıyor...", Renk.YESIL, kalin=True))
                    import subprocess as _sp
                    script_yol = os.path.abspath(__file__)
                    _sp.Popen([sys.executable, script_yol] + sys.argv[1:])
                    sys.exit(0)
                except Exception as e:
                    print(renkli(f"   ❌ Otomatik güncelleme başarısız: {e}", Renk.KIRMIZI))
                    print(renkli("   Elle güncellemek için siteden ZIP'i indirip dosyaları değiştirin.", Renk.SARI))
    test_modu = "--test" in sys.argv
    # Kullanıcı Excel dosyası verirse onu kullan, vermezse klasördeki Excel'i otomatik bul
    # Kullanım:  python sgk_bot.py [excel_dosya.xlsx]  veya  dosyayı BAT_BASLAT.bat üzerine sürükle
    verilen_dosyalar = [a for a in sys.argv[1:] if not a.startswith("--")]
    if verilen_dosyalar:
        excel_file = verilen_dosyalar[0]
    else:
        excel_file = excel_dosyasi_sec()
    try:
        if test_modu:
            # Tarayıcı açmadan arayüz akışını test eder (ilk TC'de hata simüle edilir)
            bot = SGKBot(surucu=TestSurucu(hatali_tcler=["12345678901"]), test_modu=True)
        else:
            bot = SGKBot()
        bot.run(excel_file)
    except Exception as e:
        print(renkli(f"\n❌ BOT BAŞLATILAMADI: {e}", Renk.KIRMIZI, kalin=True))
        print(renkli("   Sorunu anlamak için yukarıdaki mesajı okuyun veya BENI_OKU.txt'deki 'Sorun Giderme' bölümüne bakın.", Renk.SARI))
        input("\nKapatmak için ENTER'a basınız...")
